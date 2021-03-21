import openpyxl
import os
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

class CreateScheduleFIle:
    def __init__(self,datas,TITLE):
        self.data = datas
        self.wb = None
        self.ws = None
        self.header = ["開催日","曜日","イベントページ","開始時間"]
        self.TITLE = TITLE
    
    def MakeFormat(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        side = Side(style='thin', color='000000')
        border = Border(bottom=side)
        width = [15,7,15,15]

        # 一行目の背景色と罫線の作成
        for rows in self.ws['A1':'D1']:
            i = 0
            for cell in rows:
                cell.fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='9FD9F6',bgColor='9FD9F6')
                cell.border = border
                cell.value = self.header[i]
                cell.alignment = Alignment(horizontal='center')
                self.ws.column_dimensions[chr(i+1+64)].width = width[i]
                i = i + 1

        # 二行目から五行目の背景色
        for rows in self.ws['A2':'D5']:
            for cell in rows:
                cell.fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='D3EDFB',bgColor='D3EDFB')

    def FillInfo(self):
        i = 0
        for info in self.data:
            for j in range(4):
                self.ws.cell(row=i+2,column=j+1).value=info[j]
                self.ws.cell(row=i+2,column=j+1).alignment = Alignment(horizontal='center')
                
            self.ws.cell(row=i+2,column=3).hyperlink=info[4]
            i = i + 1
        
        self.wb.save(self.TITLE)
        os.chmod(self.TITLE,0o666)  