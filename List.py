import openpyxl
import os
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
import requests
import time
import sys
from bs4 import BeautifulSoup
from BookInfo import BookInfo

class MakeList:

    # コンストラクタ
    def __init__(self,att_name,att_id,TITLE,max_num=70):
        self.MAX_NUM = max_num # 読んだ冊数の最大値（デフォルト=50）
        self.att_name = att_name # 参加者名用配列
        self.att_id = att_id # 参加者id用配列
        self.att_num = len(att_name) # 参加人数用変数
        self.cells = [None for i in range (self.att_num * self.MAX_NUM)] # セル内項目用配列
        self.TITLE = TITLE

    # リスト用配列作成メソッド
    def MakeTable(self):
        Booknames=[] #書籍名用配列
        Bookurls=[] # 本のURL用配列

        # 参加者ごとのループ
        for i in range(self.att_num):

            # 取得URL作成、html取得
            url = 'https://bookmeter.com/users/' + self.att_id[i] + '/summary'
            res = requests.get(url)
            soup = BeautifulSoup(res.text, 'html.parser')

            print(self.att_id[i])
            # 読んだ冊数
            num = soup.select('.list__data')
            if num:
                n = int(num[0].string)
            else:
                n = 0

            print(n)

            # 読んだ冊数が制限値を超えていたらエラーを出してプログラムを停止
            if n > self.MAX_NUM:
                print("エラー：読んだ冊数が" + str(self.MAX_NUM) + "冊を超えています")
                sys.exit()

            if n != 0:
                # 読んだ本のタイトルと著者名
                Thumbnails = soup.select(".book-list--grid .book__thumbnail img")
                for Thumbnail in Thumbnails:
                    Booknames.append(Thumbnail.get('alt'))
                Authors = soup.select(".book-list--grid .detail__authors")

                Thumbnails = soup.select(".book-list--grid .thumbnail__cover a")
                for Thumbnail in Thumbnails:
                    Bookurl = Thumbnail.get('href')
                    if not Bookurl.startswith("/reviews/"):
                        Bookurls.append(Thumbnail.get('href'))

                print(Booknames)
    
            for j in range(self.MAX_NUM):
            
                # 読んだ冊数を超えた要素は空白で埋める
                if j >= n:
                    self.cells[i * self.MAX_NUM + j] = BookInfo("","","")
    
                # そうでない場合はタイトルと著者名を入れる
                else:
                    if not Authors[j].string:
                        self.cells[i * self.MAX_NUM + j] = BookInfo(Booknames[j],"",Bookurls[j])
                    else:
                        self.cells[i * self.MAX_NUM + j] = BookInfo(Booknames[j],Authors[j].string,Bookurls[j])
    
                    # 本の名前が長すぎたときに改行を入れる
                    self.cells[i * self.MAX_NUM + j].LinedTitle()
    
                    # 著者名が長過ぎたらカットする
                    self.cells[i * self.MAX_NUM + j].CutAuther()
    
                    # 本のURLを絶対パス表記に変更する
                    self.cells[i * self.MAX_NUM + j].UrlConvert()
    
            Booknames.clear()
            Bookurls.clear()
            time.sleep(10)

        return self.cells

    def MakeSpreadsheet(self):

        # エクセルシートの作成
        wb = openpyxl.Workbook()
        ws = wb.active

        # 罫線用関数の設定
        side = Side(style='thin', color='000000')
        side_dot = Side(style='dotted', color='888888')

        border1 = Border(right=side)
        border2 = Border(bottom=side_dot, right=side)
        border3 = Border(bottom=side, right=side)

        # A1セルは空白
        ws.cell(row=1,column=1,value="").border = border3

        # 一列目作成
        for i in range(self.MAX_NUM):
            ws.cell(row=3*(i+1)-1,column=1,value=i+1).alignment = Alignment(horizontal='center', vertical='top')
            ws.column_dimensions["A"].width = 3

            # 一列目の罫線描写
            ws.cell(row=3*(i+1)-1,column=1).border = border1
            ws.cell(row=3*(i+1),column=1).border = border2
            ws.cell(row=3*(i+1)+1,column=1).border = border3

        # 一行目作成
        for i in range(self.att_num):
            ws.cell(row=1,column=i+2,value=self.att_name[i]).alignment = Alignment(horizontal='center')
            ws.cell(row=1,column=i+2).border = border3

        # 本情報書き込み

        max_height = [1 for i in range(self.MAX_NUM)]  # 最大行数の初期化
        for i in range(self.att_num):
            # セル幅の設定
            ws.column_dimensions[chr(i+2+64)].width = 45
            for j in range(self.MAX_NUM):
                # 本情報の書き込みと整列
                ws.cell(row=3*(j+1)-1,column=i+2,value=self.cells[i*self.MAX_NUM+j].title).alignment = Alignment(vertical='top')
                ws.cell(row=3*(j+1)+1,column=i+2,value=self.cells[i*self.MAX_NUM+j].auther).alignment = Alignment(vertical='top')

                if self.cells[i*self.MAX_NUM+j].title != "":
                    ws.cell(row=3*(j+1),column=i+2,value="link").hyperlink = self.cells[i*self.MAX_NUM+j].url


                # 罫線描写
                ws.cell(row=3*(j+1)-1,column=i+2).border = border1
                ws.cell(row=3*(j+1),column=i+2).border = border2
                ws.cell(row=3*(j+1)+1,column=i+2).border = border3

                # 最大行数の更新
                if max_height[j] < self.cells[i*self.MAX_NUM+j].title.count("\n")+1:
                    max_height[j] = self.cells[i*self.MAX_NUM+j].title.count("\n") + 1

        # 行幅調整
        for i in range(self.MAX_NUM):
            ws.row_dimensions[3*(i+1)-1].height = 13 * max_height[i]

        # エクセルシートの保存
        wb.save(self.TITLE)
        os.chmod(self.TITLE,0o666)
    
    # テーブル作成→シート出力までの一連の処理まとめ
    def CreateSheet(self):
        self.MakeTable()
        self.MakeSpreadsheet()

