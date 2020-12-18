import json
import datetime
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
import requests
import time
import sys
import os
from bs4 import BeautifulSoup
from BookInfo import BookInfo
from Drive import DriveUpload
from GetEventInfo import EventInfo
from Zoom import ZoomAPI
from Calendar import CalendarPost
i=0

MAX_NUM = 50 # 読んだ冊数の最大値（デフォルト=50）
cells=[] # セル内項目用配列
Booknames=[] #書籍名用配列
Bookurls=[] # 本のURL用配列
OUTPUT = "output.xlsx"

# 数字をアルファベットに変換
def num2alpha(number):
    return chr(number+64)

if __name__ == "__main__":

    #input.json読み込み
    with open('input.json','r') as f:
        json_load = json.load(f)

        host = json_load["Host"]
        Event_id = json_load['EventID']
        ZoomMeetingSet = json_load['ZoomMeetingSet']
        GoogleDriveUp = json_load['GoogleDriveUp']
        GoogleCalendarUp = json_load['GoogleCalendarUp']

    event = EventInfo(Event_id,host)
    ids = event.GetMemberId()
    member_names = event.GetMemberName()
    excel_title = event.GetTitle()
    schedule = event.GetSchedule()

    print(member_names)
    print(ids)

    for i in range(len(member_names)):

    # 取得URL作成
        url = 'https://bookmeter.com/users/' + ids[i] + '/summary'
        print(url)

        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'html.parser')

        # 読んだ冊数
        num = soup.select('div.content__count')
        if num:
            n = int(num[0].string)
        else:
            n = 0

        # 読んだ冊数が制限値を超えていたらエラーを出してプログラムを停止
        if n > MAX_NUM:
            print("エラー：読んだ冊数が" + str(MAX_NUM) + "冊を超えています")
            sys.exit()

        if n != 0:
            # 読んだ本のタイトルと著者名
            Thumbnails = soup.select(".book-list--grid .book__thumbnail img")
            for Thumbnail in Thumbnails:
                Booknames.append(Thumbnail.get('alt'))
            Authors = soup.select(".book-list--grid .detail__authors")

            Thumbnails = soup.select(".book-list--grid .thumbnail__cover a")
            for Thumbnail in Thumbnails:
                Bookurl = Thumbnail.get('href')
                if not Bookurl.startswith("/reviews/"):
                    Bookurls.append(Thumbnail.get('href'))

            print(Booknames)

        for j in range(MAX_NUM):

            # 読んだ冊数を超えた要素は空白で埋める
            if j >= n:
                cells.append(BookInfo("","",""))

            # そうでない場合はタイトルと著者名を入れる
            else:
                if not Authors[j].string:
                    cells.append(BookInfo(Booknames[j],"",Bookurls[j]))
                else:
                    cells.append(BookInfo(Booknames[j],Authors[j].string,Bookurls[j]))

                # 本の名前が長すぎたときに改行を入れる
                cells[i*MAX_NUM+j].LinedTitle()

                # 著者名が長過ぎたらカットする
                cells[i*MAX_NUM+j].CutAuther()

                # 本のURLを絶対パス表記に変更する
                cells[i*MAX_NUM+j].UrlConvert()

        # Booknames.clear()
        Booknames.clear()
        Bookurls.clear()
        time.sleep(10)

    # エクセルシートの作成
    wb = openpyxl.Workbook()
    ws = wb.active

    # 罫線用関数の設定
    side = Side(style='thin', color='000000')
    side_dot = Side(style='dotted', color='888888')

    border1 = Border(right=side)
    border2 = Border(bottom=side_dot, right=side)
    border3 = Border(bottom=side, right=side)

    # A1セルは空白
    ws.cell(row=1,column=1,value="").border = border3

    # 一列目作成
    for i in range(MAX_NUM):
        ws.cell(row=3*(i+1)-1,column=1,value=i+1).alignment = Alignment(horizontal='center', vertical='top')
        ws.column_dimensions["A"].width = 3

        # 一列目の罫線描写
        ws.cell(row=3*(i+1)-1,column=1).border = border1
        ws.cell(row=3*(i+1),column=1).border = border2
        ws.cell(row=3*(i+1)+1,column=1).border = border3

    # 一行目作成
    for i in range(len(member_names)):
        ws.cell(row=1,column=i+2,value=member_names[i]).alignment = Alignment(horizontal='center')
        ws.cell(row=1,column=i+2).border = border3

    # 本情報書き込み

    max_height = [1 for i in range(MAX_NUM)]  # 最大行数の初期化
    for i in range(len(member_names)):
        # セル幅の設定
        ws.column_dimensions[num2alpha(i+2)].width = 45
        for j in range(MAX_NUM):
            # 本情報の書き込みと整列
            ws.cell(row=3*(j+1)-1,column=i+2,value=cells[i*MAX_NUM+j].title).alignment = Alignment(vertical='top')
            ws.cell(row=3*(j+1)+1,column=i+2,value=cells[i*MAX_NUM+j].auther).alignment = Alignment(vertical='top')

            if cells[i*MAX_NUM+j].title != "":
                ws.cell(row=3*(j+1),column=i+2,value="link").hyperlink = cells[i*MAX_NUM+j].url


            # 罫線描写
            ws.cell(row=3*(j+1)-1,column=i+2).border = border1
            ws.cell(row=3*(j+1),column=i+2).border = border2
            ws.cell(row=3*(j+1)+1,column=i+2).border = border3

            # 最大行数の更新
            if max_height[j] < cells[i*MAX_NUM+j].title.count("\n")+1:
                max_height[j] = cells[i*MAX_NUM+j].title.count("\n") + 1

    # 行幅調整
    for i in range(MAX_NUM):
        ws.row_dimensions[3*(i+1)-1].height = 13 * max_height[i]

    # エクセルシートの保存
    wb.save(OUTPUT)

    if(os.path.exists('client_secrets.json') and GoogleDriveUp):

        # Google Driveへのアップロード
        Drive = DriveUpload(excel_title)
        Drive.FileUpload(OUTPUT)
        D_id = Drive.GetId()
        print(D_id)
    
    # Zoomの開始時間はイベントの15分前
    d_15m = datetime.timedelta(minutes=15)
    zoom_start = schedule - d_15m
    d_120m = datetime.timedelta(minutes=120)
    zoom_end = zoom_start + d_120m

    start_time = str(zoom_start.date()) + "T" + str(zoom_start.time())
    end_time = str(zoom_end.date()) + "T" + str(zoom_end.time())
    print(start_time)

    if ZoomMeetingSet:
        zoom = ZoomAPI()
        zoom.GetUserInfo()
        zoom.CreateMeeting("LR",start_time)

    if GoogleCalendarUp:
        calendar = CalendarPost()
        calendar.PostEvent("LR","ZOOM","LR meeting",start_time,end_time)